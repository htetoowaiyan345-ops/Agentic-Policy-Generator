"""audit_diff.py

Phase 6 — write a long-format per-slot diff xlsx comparing the
previous approved version's lines_json against the new approved
version's lines_json. One row per Brain slot id (0..15).

The xlsx is written to `runs/<run_id>/audit_v<n>_diff.xlsx` so
operators have a clean per-slot diff at publish time without
having to scrape the audit_json blob.

Columns:
    run_id, version_no, slot_id, slot_label,
    before_text, after_text, changed (bool)

This module deliberately uses only `openpyxl` if present (which is
already on the requirements via `python-docx`'s sibling dependencies),
and falls back to writing CSV if openpyxl is unavailable.
"""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import List, Optional


# Mirror of the Brain slot labels used by the renderer.
_BRAIN_SLOT_LABELS = {
    0: "Free Paragraph",
    1: "Type",
    2: "Brief Description",
    3: "Approval & Effective",
    4: "Reason for Policy",
    5: "Introduction",
    6: "Policy Statement",
    7: "1. Purpose",
    8: "2. Scope & Beneficiaries",
    9: "3. Exclusions",
    10: "4. Award Structure & Payout Tiers",
    11: "5. Procedural & Compliance",
    12: "Definitions",
    13: "Related Policies, Procedures, Forms",
    14: "Policy Review Note",
    15: "History",
}


def _coerce_paragraph_text(payload) -> str:
    if isinstance(payload, dict):
        return str(payload.get('text') or '')
    if payload is None:
        return ''
    return str(payload)


def _bucket_by_slot(lines_json) -> dict:
    """Return {slot_id: '\n'.join(texts...)} for every paragraph in lines_json."""
    out: dict = {}
    if not isinstance(lines_json, list):
        return out
    for line in lines_json:
        if not isinstance(line, list) or len(line) != 2:
            continue
        kind, payload = line[0], line[1]
        if kind != 'p':
            continue
        if isinstance(payload, dict):
            slot = int(payload.get('slot', 0) or 0)
            text = str(payload.get('text') or '')
        else:
            slot = 0
            text = '' if payload is None else str(payload)
        out.setdefault(slot, []).append(text)
    return {k: "\n".join(v) for k, v in out.items()}


def _bucket_by_slot_parsed(raw_lines_json) -> dict:
    """If `raw_lines_json` arrives as a JSON string, parse it first."""
    if raw_lines_json is None:
        return {}
    if isinstance(raw_lines_json, str):
        try:
            data = json.loads(raw_lines_json)
        except Exception:
            return {}
    elif isinstance(raw_lines_json, list):
        data = raw_lines_json
    else:
        # Anything else (dict, int, …) doesn't bucket cleanly.
        return {}
    return _bucket_by_slot(data)
    return _bucket_by_slot(raw_lines_json)


def build_diff_rows(
    run_id: str,
    version_no: int,
    prev_lines_json,
    curr_lines_json,
) -> List[dict]:
    """Build one row per slot id 0..15 with the per-slot text diff."""
    prev_by_slot = _bucket_by_slot_parsed(prev_lines_json)
    curr_by_slot = _bucket_by_slot_parsed(curr_lines_json)
    rows: List[dict] = []
    for slot_id in range(0, 16):
        before = prev_by_slot.get(slot_id, '') or ''
        after = curr_by_slot.get(slot_id, '') or ''
        rows.append({
            'run_id': run_id,
            'version_no': version_no,
            'slot_id': slot_id,
            'slot_label': _BRAIN_SLOT_LABELS.get(slot_id, f'Slot {slot_id}'),
            'before_text': before,
            'after_text': after,
            'changed': (before != after),
        })
    return rows


def write_diff_xlsx(
    run_id: str,
    version_no: int,
    prev_lines_json,
    curr_lines_json,
    output_path: Path,
) -> int:
    """Write the per-slot diff xlsx (or csv fallback).

    Returns the number of rows written (always 16: slot 0..15)."""
    output_path = Path(output_path)
    rows = build_diff_rows(run_id, version_no, prev_lines_json, curr_lines_json)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    headers = list(rows[0].keys()) if rows else [
        'run_id', 'version_no', 'slot_id', 'slot_label', 'before_text', 'after_text', 'changed',
    ]

    try:
        from openpyxl import Workbook  # type: ignore
        wb = Workbook()
        ws = wb.active
        ws.title = 'SlotDiff'
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, '') for h in headers])
        wb.save(str(output_path))
        return len(rows)
    except ImportError:
        # Fallback to CSV beside the requested xlsx path (.csv extension).
        csv_path = output_path.with_suffix('.csv')
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=headers)
            w.writeheader()
            for r in rows:
                w.writerow({h: r.get(h, '') for h in headers})
        return len(rows)


def read_diff_xlsx(output_path: Path) -> List[dict]:
    """Load the diff xlsx back as a list of rows (for tests)."""
    output_path = Path(output_path)
    if not output_path.exists():
        return []
    if output_path.suffix.lower() == '.csv':
        with open(output_path, 'r', newline='', encoding='utf-8') as f:
            return list(csv.DictReader(f))
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return []
    wb = load_workbook(str(output_path))
    ws = wb.active
    rows: List[dict] = []
    headers: Optional[List[str]] = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h) if h is not None else '' for h in row]
            continue
        # Coerce None / non-str values so downstream asserts get strings.
        rows.append({
            headers[i] if i < len(headers) else f'col{i}': ('' if v is None else v)
            for i, v in enumerate(row)
        })
    return rows
